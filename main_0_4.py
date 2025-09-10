import os
import json
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import cv2
from datetime import datetime
import urllib.parse
import openpyxl
from openpyxl.styles import Font
from fpdf import FPDF
import time
import shutil
from PIL import Image, ImageTk, ImageDraw
import io
import tempfile

class LoadingWindow:
    def __init__(self, parent, title="–ü–æ–¥–æ–∂–¥–∏—Ç–µ", message="–í—ã–ø–æ–ª–Ω—è–µ—Ç—Å—è –æ–ø–µ—Ä–∞—Ü–∏—è, –ø–æ–∂–∞–ª—É–π—Å—Ç–∞, –ø–æ–¥–æ–∂–¥–∏—Ç–µ..."):
        self.parent = parent
        self.window = tk.Toplevel(parent)
        self.window.title(title)
        self.window.geometry("350x200")
        self.window.transient(parent)
        self.window.grab_set()
        self.window.resizable(False, False)
        center_window(self.window, parent)
        ico_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ico.jpg")
        try:
            pil_image = Image.open(ico_path)
            pil_image.thumbnail((80, 80), Image.Resampling.LANCZOS)
            self.photo = ImageTk.PhotoImage(pil_image)
            img_label = tk.Label(self.window, image=self.photo)
            img_label.pack(pady=10)
        except Exception as e:
            tk.Label(self.window, text="‚è≥", font=("Arial", 24)).pack(pady=10)
            print(f"–ù–µ —É–¥–∞–ª–æ—Å—å –∑–∞–≥—Ä—É–∑–∏—Ç—å ico.jpg –¥–ª—è –æ–∫–Ω–∞ –∑–∞–≥—Ä—É–∑–∫–∏: {e}")
        msg_label = tk.Label(self.window, text=message, font=("Arial", 12), wraplength=300, justify=tk.CENTER)
        msg_label.pack(pady=10)
        self.dots = 0
        self.msg_label = msg_label
        self.animate_dots()
        self.window.focus_set()

    def animate_dots(self):
        dots = "." * self.dots
        self.msg_label.config(text=f"–í—ã–ø–æ–ª–Ω—è–µ—Ç—Å—è –æ–ø–µ—Ä–∞—Ü–∏—è, –ø–æ–∂–∞–ª—É–π—Å—Ç–∞, –ø–æ–¥–æ–∂–¥–∏—Ç–µ{dots}")
        self.dots = (self.dots + 1) % 4
        self.window.after(500, self.animate_dots)

    def destroy(self):
        if self.window.winfo_exists():
            self.window.destroy()

DEFAULT_DATA_DIR = r"\\fs\share_tech\IT\py_projects\datacam"
SETTINGS_JSON = os.path.join(os.path.dirname(os.path.abspath(__file__)), "settings.json")

FONT_SIZE = 11
FONT_FAMILY = "Arial"
LARGE_FONT = (FONT_FAMILY, FONT_SIZE)
BOLD_FONT = (FONT_FAMILY, FONT_SIZE, "bold")

DEFAULT_MODELS = {
    "Hikvision DS-I200(D)": {
        "rtsp_template": "rtsp://{user}:{password}@{ip}/ISAPI/Streaming/Channels/101",
        "rtsp_template_2": "rtsp://{user}:{password}@{ip}/ISAPI/Streaming/Channels/102",
        "rtsp_template_3": "",
        "resolution": "1920x1080",
        "housing": "–ö—É–ø–æ–ª—å–Ω—ã–π, –ø–ª–∞—Å—Ç–∏–∫",
        "browser": "IE, Chrome (—Å –ø–ª–∞–≥–∏–Ω–æ–º)",
        "note": "–ë—é–¥–∂–µ—Ç–Ω–∞—è –∫–∞–º–µ—Ä–∞, H.265, –¥–µ–Ω—å/–Ω–æ—á—å, 2–ú–ø"
    },
    "Hikvision DS-I453M": {
        "rtsp_template": "rtsp://{user}:{password}@{ip}/ISAPI/Streaming/Channels/101",
        "rtsp_template_2": "rtsp://{user}:{password}@{ip}/ISAPI/Streaming/Channels/102",
        "rtsp_template_3": "rtsp://{user}:{password}@{ip}/ISAPI/Streaming/Channels/103",
        "resolution": "2688x1520",
        "housing": "–ö—É–ø–æ–ª—å–Ω—ã–π, –º–µ—Ç–∞–ª–ª",
        "browser": "IE, Chrome, Edge (Hik-Connect)",
        "note": "–ú–æ—Ç–æ—Ä–∏–∑–∏—Ä–æ–≤–∞–Ω–Ω—ã–π –æ–±—ä–µ–∫—Ç–∏–≤, 4–ú–ø, WDR, H.265"
    },
    "HiWatch DS-I400(–°) (2.8 mm)": {
        "rtsp_template": "rtsp://{user}:{password}@{ip}/ISAPI/Streaming/Channels/101",
        "rtsp_template_2": "rtsp://{user}:{password}@{ip}/ISAPI/Streaming/Channels/102",
        "rtsp_template_3": "",
        "resolution": "2560x1440",
        "housing": "–ö—É–ø–æ–ª—å–Ω—ã–π, IP67",
        "browser": "IE, Chrome (HiWatch Web)",
        "note": "–§–∏–∫—Å–∏—Ä–æ–≤–∞–Ω–Ω—ã–π –æ–±—ä–µ–∫—Ç–∏–≤ 2.8 –º–º, 4–ú–ø, IP67"
    },
    "HiWatch DS-I400(D)": {
        "rtsp_template": "rtsp://{user}:{password}@{ip}/ISAPI/Streaming/Channels/101",
        "rtsp_template_2": "rtsp://{user}:{password}@{ip}/ISAPI/Streaming/Channels/102",
        "rtsp_template_3": "",
        "resolution": "2560x1440",
        "housing": "–ö—É–ø–æ–ª—å–Ω—ã–π, –∞–Ω—Ç–∏–≤–∞–Ω–¥–∞–ª—å–Ω—ã–π",
        "browser": "IE, Chrome (HiWatch Web)",
        "note": "–î–µ–Ω—å/–Ω–æ—á—å, 4–ú–ø, –ò–ö-–ø–æ–¥—Å–≤–µ—Ç–∫–∞ 30–º"
    },
    "Dahua IPC-B040(B)": {
        "rtsp_template": "rtsp://{user}:{password}@{ip}:554/cam/realmonitor?channel=1&subtype=0",
        "rtsp_template_2": "rtsp://{user}:{password}@{ip}:554/cam/realmonitor?channel=1&subtype=1",
        "rtsp_template_3": "",
        "resolution": "2560x1440",
        "housing": "–ë—É–ª–ª–µ—Ç, –º–µ—Ç–∞–ª–ª",
        "browser": "IE, Chrome (Dahua Web)",
        "note": "4–ú–ø, Starlight, WDR, H.265, IP67"
    },
    "Novicam SMART 23 (ver.1290)": {
        "rtsp_template": "rtsp://{user}:{password}@{ip}:554/Streaming/Channels/1",
        "rtsp_template_2": "rtsp://{user}:{password}@{ip}:554/Streaming/Channels/2",
        "rtsp_template_3": "",
        "resolution": "1920x1080",
        "housing": "–ö—É–ø–æ–ª—å–Ω—ã–π, –ø–ª–∞—Å—Ç–∏–∫",
        "browser": "IE, Chrome (Novicam Web)",
        "note": "Full HD, ONVIF, P2P, –æ–±–ª–∞—á–Ω—ã–π —Å–µ—Ä–≤–∏—Å"
    },
    "AK Technology AK-IP2.4-DLV/DV28-PoE": {
        "rtsp_template": "rtsp://{user}:{password}@{ip}:554/11",
        "rtsp_template_2": "rtsp://{user}:{password}@{ip}:554/12",
        "rtsp_template_3": "",
        "resolution": "1920x1080",
        "housing": "–ö—É–ø–æ–ª—å–Ω—ã–π, –º–µ—Ç–∞–ª–ª",
        "browser": "IE, Chrome (AK Web)",
        "note": "–í–∞—Ä–∏–æ—Ñ–æ–∫–∞–ª—å–Ω—ã–π –æ–±—ä–µ–∫—Ç–∏–≤ 2.8-12 –º–º, PoE, 2–ú–ø"
    },
    "ST-V5605 PRO": {
        "rtsp_template": "rtsp://{user}:{password}@{ip}:554/stream1",
        "rtsp_template_2": "rtsp://{user}:{password}@{ip}:554/stream2",
        "rtsp_template_3": "",
        "resolution": "2560x1920",
        "housing": "–ö—É–ø–æ–ª—å–Ω—ã–π, –∞–Ω—Ç–∏–≤–∞–Ω–¥–∞–ª—å–Ω—ã–π",
        "browser": "IE, Chrome (ST Web)",
        "note": "5–ú–ø, Starlight, –ò–ö –¥–æ 40–º, H.265, IP67"
    },
    "TP-Link Tapo C200": {
        "rtsp_template": "rtsp://{user}:{password}@{ip}:554/stream1",
        "rtsp_template_2": "rtsp://{user}:{password}@{ip}:554/stream2",
        "rtsp_template_3": "",
        "resolution": "1920x1080",
        "housing": "–ü–æ–≤–æ—Ä–æ—Ç–Ω–∞—è, –ø–ª–∞—Å—Ç–∏–∫",
        "browser": "–¢–æ–ª—å–∫–æ —á–µ—Ä–µ–∑ –ø—Ä–∏–ª–æ–∂–µ–Ω–∏–µ Tapo",
        "note": "Full HD, –ø–æ–≤–æ—Ä–æ—Ç–Ω–∞—è, –æ–±–ª–∞—á–Ω—ã–π —Å–µ—Ä–≤–∏—Å, –º–∏–∫—Ä–æ—Ñ–æ–Ω/–¥–∏–Ω–∞–º–∏–∫"
    },
    "Reolink RLC-410": {
        "rtsp_template": "rtsp://{user}:{password}@{ip}:554/h264Preview_01_main",
        "rtsp_template_2": "rtsp://{user}:{password}@{ip}:554/h264Preview_01_sub",
        "rtsp_template_3": "",
        "resolution": "2560x1920",
        "housing": "–ë—É–ª–ª–µ—Ç, –º–µ—Ç–∞–ª–ª",
        "browser": "IE, Chrome (Reolink Web)",
        "note": "5–ú–ø, PoE, –ò–ö –¥–æ 30–º, ONVIF, H.265"
    }
}

def load_settings():
    if os.path.exists(SETTINGS_JSON):
        with open(SETTINGS_JSON, 'r', encoding='utf-8') as f:
            return json.load(f)
    else:
        settings = {"data_dir_path": DEFAULT_DATA_DIR}
        save_settings(settings)
        return settings

def save_settings(settings):
    with open(SETTINGS_JSON, 'w', encoding='utf-8') as f:
        json.dump(settings, f, ensure_ascii=False, indent=4)

def get_main_json_path(data_dir):
    return os.path.join(data_dir, "main.json")

def get_models_json_path(data_dir):
    return os.path.join(data_dir, "models.json")

def get_screenshots_dir(data_dir):
    return os.path.join(data_dir, "screenshots")

def ensure_data_files_exist(data_dir):
    main_path = get_main_json_path(data_dir)
    models_path = get_models_json_path(data_dir)
    created_files = []
    if not os.path.exists(main_path):
        with open(main_path, 'w', encoding='utf-8') as f:
            json.dump([], f, ensure_ascii=False, indent=4)
        created_files.append("main.json")
    if not os.path.exists(models_path):
        with open(models_path, 'w', encoding='utf-8') as f:
            json.dump(DEFAULT_MODELS, f, ensure_ascii=False, indent=4)
        created_files.append("models.json")
    screenshots_dir = get_screenshots_dir(data_dir)
    if not os.path.exists(screenshots_dir):
        os.makedirs(screenshots_dir, exist_ok=True)
        created_files.append("–ö–∞—Ç–∞–ª–æ–≥ 'screenshots'")
    return created_files

def load_models(data_dir):
    models_path = get_models_json_path(data_dir)
    if os.path.exists(models_path):
        with open(models_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    else:
        save_models(DEFAULT_MODELS, data_dir)
        return DEFAULT_MODELS

def save_models(models, data_dir):
    models_path = get_models_json_path(data_dir)
    os.makedirs(os.path.dirname(models_path), exist_ok=True)
    with open(models_path, 'w', encoding='utf-8') as f:
        json.dump(models, f, ensure_ascii=False, indent=4)

def load_cameras(data_dir):
    main_path = get_main_json_path(data_dir)
    if os.path.exists(main_path):
        try:
            with open(main_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                if isinstance(data, list):
                    return data
                else:
                    return []
        except Exception as e:
            messagebox.showerror("–û—à–∏–±–∫–∞", f"–ù–µ —É–¥–∞–ª–æ—Å—å –∑–∞–≥—Ä—É–∑–∏—Ç—å main.json: {e}")
            return []
    else:
        return []

def save_cameras(cameras, data_dir):
    main_path = get_main_json_path(data_dir)
    os.makedirs(os.path.dirname(main_path), exist_ok=True)
    with open(main_path, 'w', encoding='utf-8') as f:
        json.dump(cameras, f, ensure_ascii=False, indent=4)

def capture_rtsp_frame(rtsp_url, save_path=None):
    try:
        rtsp_url = rtsp_url.strip()
        rtsp_url = ''.join(char for char in rtsp_url if ord(char) >= 32)
        cap = cv2.VideoCapture(rtsp_url)
        cap.set(cv2.CAP_PROP_OPEN_TIMEOUT_MSEC, 5000)
        start_time = time.time()
        while not cap.isOpened():
            if time.time() - start_time > 10:
                cap.release()
                raise Exception("–¢–∞–π–º–∞—É—Ç –ø–æ–¥–∫–ª—é—á–µ–Ω–∏—è –∫ RTSP –ø–æ—Ç–æ–∫—É")
            time.sleep(0.1)
        ret, frame = cap.read()
        cap.release()
        if ret:
            if save_path:
                cv2.imwrite(save_path, frame)
            frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            pil_image = Image.fromarray(frame_rgb)
            return pil_image
        else:
            return None
    except Exception as e:
        print(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –∑–∞—Ö–≤–∞—Ç–µ –∫–∞–¥—Ä–∞: {str(e)}")
        return None

def save_camera_screenshot(rtsp_url, camera_info, data_dir):
    if not rtsp_url:
        return None
    screenshot_dir = get_screenshots_dir(data_dir)
    os.makedirs(screenshot_dir, exist_ok=True)
    safe_name = f"{camera_info.get('line', 'unknown')}_{camera_info.get('ip', 'noip')}_{camera_info.get('model', 'nomodel')}"
    safe_name = "".join(c if c.isalnum() or c in (' ', '-', '_') else '_' for c in safe_name)
    safe_name = safe_name[:100]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{safe_name}_{timestamp}.jpg"
    filepath = os.path.join(screenshot_dir, filename)
    pil_image = capture_rtsp_frame(rtsp_url)
    if pil_image:
        draw = ImageDraw.Draw(pil_image)
        try:
            from PIL import ImageFont
            font = ImageFont.load_default()
        except:
            font = None
        camera_text = f"{camera_info.get('line', '')} | {camera_info.get('model', '')} | {camera_info.get('ip', '')}"
        draw.text((5, 5), camera_text, fill="white", font=font)
        pil_image.save(filepath, "JPEG", quality=90)
        return filepath
    else:
        return None

def export_to_excel(cameras, filepath):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "–ö–∞–º–µ—Ä—ã"
    headers = ["–õ–∏–Ω–∏—è", "–ú–æ–¥–µ–ª—å", "IP", "–õ–æ–≥–∏–Ω Web", "–ü–∞—Ä–æ–ª—å Web", "–õ–æ–≥–∏–Ω –ø–æ—Ç–æ–∫", "–ü–∞—Ä–æ–ª—å –ø–æ—Ç–æ–∫", "–°–µ—Ä–∏–π–Ω—ã–π –Ω–æ–º–µ—Ä",
               "MAC", "–ü—Ä–æ—à–∏–≤–∫–∞", "–ö–æ–º–º–µ–Ω—Ç–∞—Ä–∏–π"]
    ws.append(headers)
    header_font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = header_font
    for cam in cameras:
        row = [
            cam.get("line", ""),
            cam.get("model", ""),
            cam.get("ip", ""),
            cam.get("web_user", ""),
            cam.get("web_pass", ""),
            cam.get("stream_user", ""),
            cam.get("stream_pass", ""),
            cam.get("sn", ""),
            cam.get("mac", ""),
            cam.get("fw", ""),
            cam.get("comment", "")
        ]
        ws.append(row)
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    wb.save(filepath)
    messagebox.showinfo("–≠–∫—Å–ø–æ—Ä—Ç", f"–î–∞–Ω–Ω—ã–µ —É—Å–ø–µ—à–Ω–æ —ç–∫—Å–ø–æ—Ä—Ç–∏—Ä–æ–≤–∞–Ω—ã –≤ Excel: {filepath}")

def export_to_pdf(cameras, data_dir, filepath):
    pdf = FPDF(orientation='L')
    pdf.add_page()
    font_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ChakraPetch-Regular.ttf")
    if os.path.exists(font_path):
        try:
            pdf.add_font("ChakraPetch", "", font_path, uni=True)
            pdf.set_font("ChakraPetch", size=8)
        except Exception as e:
            pdf.set_font("Arial", size=8)
            messagebox.showwarning("–®—Ä–∏—Ñ—Ç", f"–ù–µ —É–¥–∞–ª–æ—Å—å –∑–∞–≥—Ä—É–∑–∏—Ç—å ChakraPetch-Regular.ttf: {str(e)} ‚Äî –∫–∏—Ä–∏–ª–ª–∏—Ü–∞ –º–æ–∂–µ—Ç –æ—Ç–æ–±—Ä–∞–∂–∞—Ç—å—Å—è –Ω–µ–∫–æ—Ä—Ä–µ–∫—Ç–Ω–æ.")
    else:
        pdf.set_font("Arial", size=8)
        messagebox.showwarning("–®—Ä–∏—Ñ—Ç", "–ù–µ –Ω–∞–π–¥–µ–Ω ChakraPetch-Regular.ttf ‚Äî –∫–∏—Ä–∏–ª–ª–∏—Ü–∞ –º–æ–∂–µ—Ç –æ—Ç–æ–±—Ä–∞–∂–∞—Ç—å—Å—è –Ω–µ–∫–æ—Ä—Ä–µ–∫—Ç–Ω–æ.")
    pdf.cell(0, 10, txt="–î–µ—Ç–∞–ª—å–Ω–æ–µ –¥–æ—Å—å–µ IP –∫–∞–º–µ—Ä", ln=True, align='C')
    pdf.ln(10)
    col_widths = {
        "screenshot": 25,
        "line": 25,
        "model": 35,
        "ip": 25,
        "web_user": 20,
        "web_pass": 20,
        "stream_user": 20,
        "stream_pass": 20,
        "mac": 32,
        "comment": 55
    }
    headers = [
        ("–°–∫—Ä–∏–Ω", col_widths["screenshot"]),
        ("–õ–∏–Ω–∏—è", col_widths["line"]),
        ("–ú–æ–¥–µ–ª—å", col_widths["model"]),
        ("IP", col_widths["ip"]),
        ("Web –ª–æ–≥–∏–Ω", col_widths["web_user"]),
        ("Web –ø–∞—Ä–æ–ª—å", col_widths["web_pass"]),
        ("–ü–æ—Ç–æ–∫ –ª–æ–≥–∏–Ω", col_widths["stream_user"]),
        ("–ü–æ—Ç–æ–∫ –ø–∞—Ä–æ–ª—å", col_widths["stream_pass"]),
        ("MAC-–∞–¥—Ä–µ—Å", col_widths["mac"]),
        ("–ö–æ–º–º–µ–Ω—Ç–∞—Ä–∏–π", col_widths["comment"])
    ]
    for header, width in headers:
        pdf.cell(width, 10, header, border=1, align='C')
    pdf.ln()
    models = load_models(data_dir)
    for cam in cameras:
        row_data = [
            ("", col_widths["screenshot"]),
            (cam.get("line", "")[:20], col_widths["line"]),
            (cam.get("model", "")[:30], col_widths["model"]),
            (cam.get("ip", "")[:15], col_widths["ip"]),
            (cam.get("web_user", "")[:12], col_widths["web_user"]),
            (cam.get("web_pass", "")[:12], col_widths["web_pass"]),
            (cam.get("stream_user", "")[:12], col_widths["stream_user"]),
            (cam.get("stream_pass", "")[:12], col_widths["stream_pass"]),
            (cam.get("mac", "")[:20], col_widths["mac"]),
            (cam.get("comment", "")[:40], col_widths["comment"])
        ]
        rtsp_url = ""
        model_name = cam.get("model", "")
        model_data = models.get(model_name, {})
        rtsp_template = model_data.get("rtsp_template", "")
        if rtsp_template and cam.get("ip"):
            user = urllib.parse.quote(cam.get("stream_user", "admin"), safe='@$')
            password = urllib.parse.quote(cam.get("stream_pass", "admin"), safe='@$')
            ip = cam.get("ip", "")
            rtsp_url = rtsp_template.format(user=user, password=password, ip=ip)
        pil_image = None
        if rtsp_url:
            pil_image = capture_rtsp_frame(rtsp_url)
        if pil_image:
            pil_image.thumbnail((64, 36), Image.Resampling.LANCZOS)
            if pil_image.mode in ("RGBA", "P"):
                pil_image = pil_image.convert("RGB")
            with tempfile.NamedTemporaryFile(delete=False, suffix=".jpg") as tmp_file:
                temp_img_path = tmp_file.name
            pil_image.save(temp_img_path, "JPEG", quality=85)
            try:
                pdf.image(temp_img_path, x=pdf.get_x()+2, y=pdf.get_y()+1, w=21, h=8)
                pdf.cell(col_widths["screenshot"], 10, "", border=1)
            except Exception as img_e:
                print(f"–û—à–∏–±–∫–∞ –≤—Å—Ç–∞–≤–∫–∏ –∏–∑–æ–±—Ä–∞–∂–µ–Ω–∏—è: {img_e}")
                pdf.cell(col_widths["screenshot"], 10, "–ù/–î", border=1, align='C')
            finally:
                if os.path.exists(temp_img_path):
                    os.remove(temp_img_path)
        else:
            pdf.cell(col_widths["screenshot"], 10, "–ù/–î", border=1, align='C')
        for i, (text, width) in enumerate(row_data[1:], start=1):
            pdf.cell(width, 10, text, border=1)
        pdf.ln()
    pdf.output(filepath)
    messagebox.showinfo("–≠–∫—Å–ø–æ—Ä—Ç", f"–ü–æ–¥—Ä–æ–±–Ω–æ–µ –¥–æ—Å—å–µ —É—Å–ø–µ—à–Ω–æ —ç–∫—Å–ø–æ—Ä—Ç–∏—Ä–æ–≤–∞–Ω–æ –≤ PDF: {filepath}")

def generate_gallery(cameras, data_dir, filepath):
    if not cameras:
        messagebox.showwarning("–ì–∞–ª–µ—Ä–µ—è", "–ù–µ—Ç –∫–∞–º–µ—Ä –¥–ª—è —Å–æ–∑–¥–∞–Ω–∏—è –≥–∞–ª–µ—Ä–µ–∏.")
        return
    models = load_models(data_dir)
    screenshots = []
    for cam in cameras:
        rtsp_url = ""
        model_name = cam.get("model", "")
        model_data = models.get(model_name, {})
        rtsp_template = model_data.get("rtsp_template", "")
        if rtsp_template and cam.get("ip"):
            user = urllib.parse.quote(cam.get("stream_user", "admin"), safe='@$')
            password = urllib.parse.quote(cam.get("stream_pass", "admin"), safe='@$')
            ip = cam.get("ip", "")
            rtsp_url = rtsp_template.format(user=user, password=password, ip=ip)
        if rtsp_url:
            saved_path = save_camera_screenshot(rtsp_url, cam, data_dir)
            if saved_path:
                try:
                    pil_image = Image.open(saved_path)
                    pil_image.thumbnail((320, 180), Image.Resampling.LANCZOS)
                    screenshots.append(pil_image)
                except Exception as e:
                    print(f"–û—à–∏–±–∫–∞ –∑–∞–≥—Ä—É–∑–∫–∏ —Å–∫—Ä–∏–Ω—à–æ—Ç–∞ {saved_path}: {e}")
    if not screenshots:
        messagebox.showwarning("–ì–∞–ª–µ—Ä–µ—è", "–ù–µ —É–¥–∞–ª–æ—Å—å –ø–æ–ª—É—á–∏—Ç—å –Ω–∏ –æ–¥–Ω–æ–≥–æ —Å–∫—Ä–∏–Ω—à–æ—Ç–∞.")
        return
    images_per_row = 3
    rows = (len(screenshots) + images_per_row - 1) // images_per_row
    cell_width, cell_height = 320, 180
    padding = 10
    gallery_width = images_per_row * cell_width + (images_per_row + 1) * padding
    gallery_height = rows * cell_height + (rows + 1) * padding
    gallery_image = Image.new('RGB', (gallery_width, gallery_height), color='black')
    for idx, img in enumerate(screenshots):
        row = idx // images_per_row
        col = idx % images_per_row
        x = padding + col * (cell_width + padding)
        y = padding + row * (cell_height + padding)
        gallery_image.paste(img, (x, y))
    try:
        gallery_image.save(filepath)
        screenshot_dir = get_screenshots_dir(data_dir)
        messagebox.showinfo("–ì–∞–ª–µ—Ä–µ—è", f"–ì–∞–ª–µ—Ä–µ—è —É—Å–ø–µ—à–Ω–æ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∞: {filepath}\n–°–∫—Ä–∏–Ω—à–æ—Ç—ã –∫–∞–º–µ—Ä —Å–æ—Ö—Ä–∞–Ω–µ–Ω—ã –≤: {screenshot_dir}")
    except Exception as e:
        messagebox.showerror("–û—à–∏–±–∫–∞", f"–ù–µ —É–¥–∞–ª–æ—Å—å —Å–æ—Ö—Ä–∞–Ω–∏—Ç—å –≥–∞–ª–µ—Ä–µ—é: {str(e)}")

def center_window(win, parent):
    win.update_idletasks()
    width = win.winfo_width()
    height = win.winfo_height()
    x = parent.winfo_x() + (parent.winfo_width() // 2) - (width // 2)
    y = parent.winfo_y() + (parent.winfo_height() // 2) - (height // 2)
    win.geometry(f"+{x}+{y}")

class ModelSelectorWindow:
    def __init__(self, parent, current_value, models_list, callback):
        self.parent = parent
        self.callback = callback
        self.window = tk.Toplevel(parent)
        self.window.title("–í—ã–±–æ—Ä –º–æ–¥–µ–ª–∏")
        self.window.geometry("400x200")
        center_window(self.window, parent)
        self.window.transient(parent)
        self.window.grab_set()
        tk.Label(self.window, text="–í—ã–±–µ—Ä–∏—Ç–µ –º–æ–¥–µ–ª—å:", font=LARGE_FONT).pack(pady=10)
        self.model_var = tk.StringVar(value=current_value)
        self.model_combo = ttk.Combobox(self.window, textvariable=self.model_var, values=models_list, state="readonly", font=LARGE_FONT, width=45)
        self.model_combo.pack(pady=5, padx=20)
        btn_frame = tk.Frame(self.window)
        btn_frame.pack(pady=20)
        tk.Button(btn_frame, text="üíæ –í—ã–±—Ä–∞—Ç—å", command=self.save_selection, font=LARGE_FONT, bg="#a0f0a0").pack(side=tk.LEFT, padx=10)
        tk.Button(btn_frame, text="‚ùå –û—Ç–º–µ–Ω–∞", command=self.window.destroy, font=LARGE_FONT).pack(side=tk.LEFT, padx=10)

    def save_selection(self):
        selected_model = self.model_var.get().strip()
        if selected_model:
            self.callback(selected_model)
        self.window.destroy()

class RTSPDetailWindow:
    def __init__(self, parent, camera_index, cameras, models, app_instance, data_dir):
        self.parent = parent
        self.camera_index = camera_index
        self.cameras = cameras
        self.models = models
        self.app = app_instance
        self.data_dir = data_dir
        self.camera = cameras[camera_index].copy()
        self.rtsp_url = ""
        self.screenshot_label = None
        self.photo = None
        self.current_pil_image = None
        self.use_web_creds = tk.BooleanVar(value=False)
        self.window = tk.Toplevel(parent)
        self.window.title(f"–î–µ—Ç–∞–ª–∏ –∫–∞–º–µ—Ä—ã: {self.camera.get('model', '–ù–µ–∏–∑–≤–µ—Å—Ç–Ω–æ')} ({self.camera.get('ip', '–ù–µ—Ç IP')})")
        self.window.geometry("900x850")
        center_window(self.window, parent)
        self.window.transient(parent)
        self.window.grab_set()

        canvas = tk.Canvas(self.window)
        scrollbar = ttk.Scrollbar(self.window, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        fields = [
            ("–õ–∏–Ω–∏—è", "line"),
            ("IP-–∞–¥—Ä–µ—Å", "ip"),
            ("–õ–æ–≥–∏–Ω Web", "web_user"),
            ("–ü–∞—Ä–æ–ª—å Web", "web_pass"),
            ("–õ–æ–≥–∏–Ω –ø–æ—Ç–æ–∫–∞", "stream_user"),
            ("–ü–∞—Ä–æ–ª—å –ø–æ—Ç–æ–∫–∞", "stream_pass"),
            ("–°–µ—Ä–∏–π–Ω—ã–π –Ω–æ–º–µ—Ä", "sn"),
            ("MAC-–∞–¥—Ä–µ—Å", "mac"),
            ("–ü—Ä–æ—à–∏–≤–∫–∞", "fw"),
            ("–ö–æ–º–º–µ–Ω—Ç–∞—Ä–∏–π", "comment")
        ]
        self.entries = {}
        row = 0
        for i, (label_text, key) in enumerate(fields):
            tk.Label(scrollable_frame, text=f"{label_text}:", font=BOLD_FONT, anchor="w").grid(row=row, column=0, sticky="w", pady=5, padx=10)
            var = tk.StringVar(value=self.camera.get(key, ""))
            entry = tk.Entry(scrollable_frame, textvariable=var, width=60, font=LARGE_FONT)
            entry.grid(row=row, column=1, sticky="w", padx=10, pady=5)
            self.entries[key] = var
            # –£–Ω–∏–≤–µ—Ä—Å–∞–ª—å–Ω–∞—è –ø—Ä–∏–≤—è–∑–∫–∞ –¥–ª—è Ctrl+C/V –Ω–∞ —Ä—É—Å—Å–∫–æ–π –∏ –∞–Ω–≥–ª–∏–π—Å–∫–æ–π —Ä–∞—Å–∫–ª–∞–¥–∫–µ
            def handle_ctrl_c_v(event, entry_ref=entry):
                if event.state & 0x4:  # Ctrl –Ω–∞–∂–∞—Ç
                    if event.keycode == 54:  # –ö–ª–∞–≤–∏—à–∞ 'C' / '–°'
                        entry_ref.event_generate("<<Copy>>")
                        return "break"
                    elif event.keycode == 58:  # –ö–ª–∞–≤–∏—à–∞ 'V' / '–ú'
                        entry_ref.event_generate("<<Paste>>")
                        return "break"
            entry.bind("<Key>", handle_ctrl_c_v)
            row += 1

        tk.Label(scrollable_frame, text="–ú–æ–¥–µ–ª—å:", font=BOLD_FONT, anchor="w").grid(row=row, column=0, sticky="w", pady=5, padx=10)
        self.model_var = tk.StringVar(value=self.camera.get("model", ""))
        self.model_combo = ttk.Combobox(scrollable_frame, textvariable=self.model_var, values=list(self.models.keys()), state="readonly", width=58, font=LARGE_FONT)
        self.model_combo.grid(row=row, column=1, sticky="w", padx=10, pady=5)
        row += 1

        model_name = self.camera.get("model", "")
        model_data = models.get(model_name, {})
        rtsp_template = model_data.get("rtsp_template", "")
        if rtsp_template and self.camera.get("ip"):
            if self.use_web_creds.get():
                user = urllib.parse.quote(self.camera.get("web_user", "admin"), safe='@$')
                password = urllib.parse.quote(self.camera.get("web_pass", "admin"), safe='@$')
            else:
                user = urllib.parse.quote(self.camera.get("stream_user", "admin"), safe='@$')
                password = urllib.parse.quote(self.camera.get("stream_pass", "admin"), safe='@$')
            ip = self.camera.get("ip", "")
            self.rtsp_url = rtsp_template.format(user=user, password=password, ip=ip)
        else:
            self.rtsp_url = "–ù–µ —É–¥–∞–ª–æ—Å—å —Å—Ñ–æ—Ä–º–∏—Ä–æ–≤–∞—Ç—å RTSP-—Å—Å—ã–ª–∫—É"

        self.use_web_check = tk.Checkbutton(scrollable_frame, text="–ò—Å–ø–æ–ª—å–∑–æ–≤–∞—Ç—å Web –ª–æ–≥–∏–Ω/–ø–∞—Ä–æ–ª—å", variable=self.use_web_creds, font=LARGE_FONT, command=self.update_rtsp_url)
        self.use_web_check.grid(row=row, column=0, columnspan=2, sticky="w", padx=10, pady=5)
        row += 1

        tk.Label(scrollable_frame, text="RTSP URL (–æ—Å–Ω–æ–≤–Ω–æ–π):", font=BOLD_FONT, anchor="w").grid(row=row, column=0, sticky="w", pady=5, padx=10)
        self.url_label = tk.Label(scrollable_frame, text=self.rtsp_url, font=("Courier", 10), fg="blue", anchor="w", wraplength=500, justify="left")
        self.url_label.grid(row=row, column=1, sticky="w", padx=10, pady=5)
        row += 1

        btn_frame_top = tk.Frame(scrollable_frame, pady=10)
        btn_frame_top.grid(row=row, column=0, columnspan=3, pady=10)
        tk.Button(btn_frame_top, text="üíæ –°–æ—Ö—Ä–∞–Ω–∏—Ç—å –∏–∑–º–µ–Ω–µ–Ω–∏—è", command=self.save_changes, font=LARGE_FONT, bg="#a0f0a0").pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame_top, text="üì∏ –°–¥–µ–ª–∞—Ç—å —Å–∫—Ä–∏–Ω—à–æ—Ç", command=self.load_screenshot, font=LARGE_FONT, bg="#a0d0f0").pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame_top, text="üìã –ö–æ–ø–∏—Ä–æ–≤–∞—Ç—å RTSP URL", command=self.copy_rtsp_url, font=LARGE_FONT).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame_top, text="‚ùå –ó–∞–∫—Ä—ã—Ç—å", command=self.window.destroy, font=LARGE_FONT).pack(side=tk.LEFT, padx=5)
        row += 1

        tk.Label(scrollable_frame, text="–°–∫—Ä–∏–Ω—à–æ—Ç —Å –∫–∞–º–µ—Ä—ã:", font=BOLD_FONT).grid(row=row, column=0, columnspan=2, sticky="w", pady=(10,5), padx=10)
        row += 1

        self.screenshot_frame = tk.Frame(scrollable_frame, width=720, height=405, relief="sunken", bd=1)
        self.screenshot_frame.grid(row=row, column=0, columnspan=2, padx=10, pady=5)
        self.screenshot_frame.grid_propagate(False)
        self.screenshot_label = tk.Label(self.screenshot_frame, text="–ù–∞–∂–º–∏—Ç–µ '–°–¥–µ–ª–∞—Ç—å —Å–∫—Ä–∏–Ω—à–æ—Ç'", font=LARGE_FONT)
        self.screenshot_label.place(relx=0.5, rely=0.5, anchor="center")
        row += 1

        # –ö–Ω–æ–ø–∫–∞ –¥–ª—è —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏—è —Ç–µ–∫—É—â–µ–≥–æ —Å–∫—Ä–∏–Ω—à–æ—Ç–∞
        tk.Button(scrollable_frame, text="üíæ –°–æ—Ö—Ä–∞–Ω–∏—Ç—å —Å–∫—Ä–∏–Ω—à–æ—Ç", command=self.save_current_screenshot, font=LARGE_FONT, bg="#f0d0a0").grid(row=row, column=0, columnspan=2, pady=5)
        row += 1

        btn_frame_bottom = tk.Frame(scrollable_frame, pady=20)
        btn_frame_bottom.grid(row=row, column=0, columnspan=2, pady=20)
        tk.Button(btn_frame_bottom, text="üíæ –°–æ—Ö—Ä–∞–Ω–∏—Ç—å –∏–∑–º–µ–Ω–µ–Ω–∏—è", command=self.save_changes, font=LARGE_FONT, bg="#a0f0a0").pack(side=tk.LEFT, padx=10)
        tk.Button(btn_frame_bottom, text="‚ùå –ó–∞–∫—Ä—ã—Ç—å", command=self.window.destroy, font=LARGE_FONT).pack(side=tk.LEFT, padx=10)

    def update_rtsp_url(self):
        model_name = self.model_var.get()
        model_data = self.models.get(model_name, {})
        rtsp_template = model_data.get("rtsp_template", "")
        if not rtsp_template or not self.camera.get("ip"):
            self.rtsp_url = "–ù–µ —É–¥–∞–ª–æ—Å—å —Å—Ñ–æ—Ä–º–∏—Ä–æ–≤–∞—Ç—å RTSP-—Å—Å—ã–ª–∫—É"
        else:
            if self.use_web_creds.get():
                user = urllib.parse.quote(self.camera.get("web_user", "admin"), safe='@$')
                password = urllib.parse.quote(self.camera.get("web_pass", "admin"), safe='@$')
            else:
                user = urllib.parse.quote(self.camera.get("stream_user", "admin"), safe='@$')
                password = urllib.parse.quote(self.camera.get("stream_pass", "admin"), safe='@$')
            ip = self.camera.get("ip", "")
            self.rtsp_url = rtsp_template.format(user=user, password=password, ip=ip)
        if hasattr(self, 'url_label'):
            self.url_label.config(text=self.rtsp_url)

    def copy_rtsp_url(self):
        if self.rtsp_url and "–ù–µ —É–¥–∞–ª–æ—Å—å" not in self.rtsp_url:
            self.window.clipboard_clear()
            self.window.clipboard_append(self.rtsp_url)
            self.window.update()
            messagebox.showinfo("–ö–æ–ø–∏—Ä–æ–≤–∞–Ω–∏–µ", "RTSP URL —Å–∫–æ–ø–∏—Ä–æ–≤–∞–Ω –≤ –±—É—Ñ–µ—Ä –æ–±–º–µ–Ω–∞!")
        else:
            messagebox.showwarning("–û—à–∏–±–∫–∞", "–ù–µ–≤–æ–∑–º–æ–∂–Ω–æ —Å–∫–æ–ø–∏—Ä–æ–≤–∞—Ç—å: RTSP URL –Ω–µ —Å—Ñ–æ—Ä–º–∏—Ä–æ–≤–∞–Ω.")

    def load_screenshot(self):
        model_name = self.model_var.get()
        model_data = self.models.get(model_name, {})
        rtsp_template = model_data.get("rtsp_template", "")
        if not rtsp_template or not self.camera.get("ip"):
            if self.screenshot_label and self.screenshot_label.winfo_exists():
                self.screenshot_label.config(text="‚ùå –ù–µ —É–¥–∞–ª–æ—Å—å –∑–∞–≥—Ä—É–∑–∏—Ç—å —Å–∫—Ä–∏–Ω—à–æ—Ç: –Ω–µ—Ç —à–∞–±–ª–æ–Ω–∞ RTSP –∏–ª–∏ IP")
            return

        if self.use_web_creds.get():
            user = urllib.parse.quote(self.camera.get("web_user", "admin"), safe='@$')
            password = urllib.parse.quote(self.camera.get("web_pass", "admin"), safe='@$')
        else:
            user = urllib.parse.quote(self.camera.get("stream_user", "admin"), safe='@$')
            password = urllib.parse.quote(self.camera.get("stream_pass", "admin"), safe='@$')
        ip = self.camera.get("ip", "")
        rtsp_url_for_screenshot = rtsp_template.format(user=user, password=password, ip=ip)

        if self.screenshot_label and self.screenshot_label.winfo_exists():
            self.screenshot_label.config(text="–ó–∞–≥—Ä—É–∑–∫–∞ —Å–∫—Ä–∏–Ω—à–æ—Ç–∞...")

        pil_image = capture_rtsp_frame(rtsp_url_for_screenshot)
        if pil_image:
            save_camera_screenshot(rtsp_url_for_screenshot, self.camera, self.data_dir)
            self.current_pil_image = pil_image.copy()
            pil_image.thumbnail((720, 405))
            self.photo = ImageTk.PhotoImage(pil_image)
            if self.screenshot_label and self.screenshot_label.winfo_exists():
                self.screenshot_label.config(image=self.photo, text="")
        else:
            if self.screenshot_label and self.screenshot_label.winfo_exists():
                self.screenshot_label.config(text="‚ùå –ù–µ —É–¥–∞–ª–æ—Å—å –ø–æ–ª—É—á–∏—Ç—å –∫–∞–¥—Ä —Å –∫–∞–º–µ—Ä—ã")

    def save_current_screenshot(self):
        if not hasattr(self, 'photo') or self.photo is None:
            messagebox.showwarning("–û—à–∏–±–∫–∞", "–ù–µ—Ç —Å–∫—Ä–∏–Ω—à–æ—Ç–∞ –¥–ª—è —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏—è. –°–Ω–∞—á–∞–ª–∞ —Å–¥–µ–ª–∞–π—Ç–µ —Å–∫—Ä–∏–Ω—à–æ—Ç.")
            return

        filepath = filedialog.asksaveasfilename(
            defaultextension=".jpg",
            filetypes=[("JPEG files", "*.jpg"), ("PNG files", "*.png"), ("All files", "*.*")],
            title="–°–æ—Ö—Ä–∞–Ω–∏—Ç—å —Å–∫—Ä–∏–Ω—à–æ—Ç"
        )

        if not filepath:
            return

        try:
            if not hasattr(self, 'current_pil_image') or self.current_pil_image is None:
                raise Exception("–ò—Å—Ö–æ–¥–Ω–æ–µ –∏–∑–æ–±—Ä–∞–∂–µ–Ω–∏–µ –Ω–µ –Ω–∞–π–¥–µ–Ω–æ.")

            if filepath.lower().endswith('.png'):
                self.current_pil_image.save(filepath, "PNG")
            else:
                self.current_pil_image.save(filepath, "JPEG", quality=90)

            messagebox.showinfo("–£—Å–ø–µ—Ö", f"–°–∫—Ä–∏–Ω—à–æ—Ç —É—Å–ø–µ—à–Ω–æ —Å–æ—Ö—Ä–∞–Ω–µ–Ω:\n{filepath}")

        except Exception as e:
            messagebox.showerror("–û—à–∏–±–∫–∞", f"–ù–µ —É–¥–∞–ª–æ—Å—å —Å–æ—Ö—Ä–∞–Ω–∏—Ç—å —Å–∫—Ä–∏–Ω—à–æ—Ç:\n{str(e)}")

    def save_changes(self):
        for key, var in self.entries.items():
            self.camera[key] = var.get().strip()
        self.camera["model"] = self.model_var.get().strip()
        self.cameras[self.camera_index] = self.camera
        save_cameras(self.cameras, self.data_dir)
        self.app.refresh_table()
        messagebox.showinfo("–£—Å–ø–µ—Ö", "–î–∞–Ω–Ω—ã–µ –∫–∞–º–µ—Ä—ã —Å–æ—Ö—Ä–∞–Ω–µ–Ω—ã")
        self.window.destroy()

class CameraApp:
    def __init__(self, root):
        self.root = root
        self.root.title("IP Camera Manager")
        self.root.state('zoomed')
        self.style = ttk.Style()
        self.style.configure("Treeview", font=LARGE_FONT, rowheight=25)
        self.style.configure("Treeview.Heading", font=BOLD_FONT)
        self.settings = load_settings()
        self.data_dir = self.settings["data_dir_path"]
        self.models = load_models(self.data_dir)
        self.cameras = load_cameras(self.data_dir)
        self.filtered_cameras = self.cameras.copy()
        self.rtsp_buttons = []

        top_frame = tk.Frame(root, pady=5, padx=10, relief="groove", bd=1)
        top_frame.pack(fill=tk.X)
        tk.Button(top_frame, text="‚ûï –î–æ–±–∞–≤–∏—Ç—å –∫–∞–º–µ—Ä—É", command=self.open_add_camera_tab, font=LARGE_FONT, bg="#a0f0a0", relief="groove").pack(side=tk.LEFT, padx=5)
        tk.Button(top_frame, text="üìã –°–ø–∏—Å–æ–∫ –∫–∞–º–µ—Ä", command=self.show_camera_list, font=LARGE_FONT, relief="groove").pack(side=tk.LEFT, padx=5)
        tk.Button(top_frame, text="üîÑ –û–±–Ω–æ–≤–∏—Ç—å", command=self.refresh_table, font=LARGE_FONT, relief="groove").pack(side=tk.LEFT, padx=5)
        tk.Button(top_frame, text="üíæ –°–¥–µ–ª–∞—Ç—å —Ä–µ–∑–µ—Ä–≤", command=self.backup_main_json, font=LARGE_FONT, bg="#f0f0a0", relief="groove").pack(side=tk.LEFT, padx=5)
        tk.Button(top_frame, text="‚öôÔ∏è –ù–∞—Å—Ç—Ä–æ–π–∫–∏", command=self.open_settings_tab, font=LARGE_FONT, relief="groove").pack(side=tk.LEFT, padx=5)
        tk.Button(top_frame, text="üìã –ú–æ–¥–µ–ª–∏", command=self.open_models_tab, font=LARGE_FONT, relief="groove").pack(side=tk.LEFT, padx=5)
        tk.Button(top_frame, text="üìÑ –í—ã–≥—Ä—É–∑–∫–∞ –≤ PDF", command=self.export_pdf, font=LARGE_FONT, relief="groove").pack(side=tk.LEFT, padx=5)
        tk.Button(top_frame, text="üìä –≠–∫—Å–ø–æ—Ä—Ç –≤ Excel", command=self.export_excel, font=LARGE_FONT, relief="groove").pack(side=tk.LEFT, padx=5)
        tk.Button(top_frame, text="üñºÔ∏è –ì–∞–ª–µ—Ä–µ—è", command=self.generate_gallery, font=LARGE_FONT, relief="groove", bg="#f0d0a0").pack(side=tk.LEFT, padx=5)
        # ‚úÖ –ö–ù–û–ü–ö–ê "–°–∫—Ä–∏–Ω—à–æ—Ç—ã –≤—Å–µ—Ö" –£–î–ê–õ–ï–ù–ê
        # tk.Button(top_frame, text="üì∏ –°–∫—Ä–∏–Ω—à–æ—Ç—ã –≤—Å–µ—Ö", command=self.take_all_screenshots, font=LARGE_FONT, relief="groove", bg="#d0a0f0").pack(side=tk.LEFT, padx=5)
        tk.Button(top_frame, text="‚ÑπÔ∏è –ò–Ω—Ñ–æ", command=self.show_info, font=LARGE_FONT, relief="groove").pack(side=tk.RIGHT, padx=5)

        search_frame = tk.Frame(root, pady=10, padx=10)
        search_frame.pack(fill=tk.X)
        tk.Label(search_frame, text="üîç –ü–æ–∏—Å–∫:", font=LARGE_FONT).pack(side=tk.LEFT, padx=(0, 5))
        self.search_var = tk.StringVar()
        self.search_var.trace("w", self.on_search_change)
        search_entry = tk.Entry(search_frame, textvariable=self.search_var, width=50, font=LARGE_FONT)
        search_entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        self.notebook = ttk.Notebook(root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        self.list_frame = tk.Frame(self.notebook)
        self.notebook.add(self.list_frame, text="–°–ø–∏—Å–æ–∫ –∫–∞–º–µ—Ä")

        table_frame = tk.Frame(self.list_frame)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        columns = ("line", "model", "ip", "web_user", "web_pass", "stream_user", "stream_pass", "sn", "mac", "fw", "comment", "rtsp_btn")
        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings", selectmode="browse")
        self.tree.heading("line", text="–õ–∏–Ω–∏—è", command=lambda: self.sort_column("line", False))
        self.tree.heading("model", text="–ú–æ–¥–µ–ª—å", command=lambda: self.sort_column("model", False))
        self.tree.heading("ip", text="IP", command=lambda: self.sort_column("ip", False))
        self.tree.heading("web_user", text="Web –ª–æ–≥–∏–Ω", command=lambda: self.sort_column("web_user", False))
        self.tree.heading("web_pass", text="Web –ø–∞—Ä–æ–ª—å", command=lambda: self.sort_column("web_pass", False))
        self.tree.heading("stream_user", text="–ü–æ—Ç–æ–∫ –ª–æ–≥–∏–Ω", command=lambda: self.sort_column("stream_user", False))
        self.tree.heading("stream_pass", text="–ü–æ—Ç–æ–∫ –ø–∞—Ä–æ–ª—å", command=lambda: self.sort_column("stream_pass", False))
        self.tree.heading("sn", text="–°–µ—Ä–∏–π–Ω—ã–π –Ω–æ–º–µ—Ä", command=lambda: self.sort_column("sn", False))
        self.tree.heading("mac", text="MAC", command=lambda: self.sort_column("mac", False))
        self.tree.heading("fw", text="–ü—Ä–æ—à–∏–≤–∫–∞", command=lambda: self.sort_column("fw", False))
        self.tree.heading("comment", text="–ö–æ–º–º–µ–Ω—Ç–∞—Ä–∏–π", command=lambda: self.sort_column("comment", False))
        self.tree.heading("rtsp_btn", text="")

        self.tree.column("line", width=120, anchor="w")
        self.tree.column("model", width=180, anchor="w")
        self.tree.column("ip", width=120, anchor="w")
        self.tree.column("web_user", width=100, anchor="w")
        self.tree.column("web_pass", width=100, anchor="w")
        self.tree.column("stream_user", width=100, anchor="w")
        self.tree.column("stream_pass", width=100, anchor="w")
        self.tree.column("sn", width=150, anchor="w")
        self.tree.column("mac", width=150, anchor="w")
        self.tree.column("fw", width=120, anchor="w")
        self.tree.column("comment", width=200, anchor="w")
        self.tree.column("rtsp_btn", width=120, anchor="center")

        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        self.tree.bind("<Double-1>", self.on_double_click_edit)
        self.tree.bind("<Button-3>", self.on_right_click)
        self.tree.bind("<Configure>", self.schedule_button_placement)
        self.tree.bind("<MouseWheel>", self.schedule_button_placement)
        self.tree.bind("<Button-4>", self.schedule_button_placement)
        self.tree.bind("<Button-5>", self.schedule_button_placement)

        self.refresh_table()

        self.status_bar = tk.Label(root, text="", bd=1, relief=tk.SUNKEN, anchor=tk.W, font=LARGE_FONT)
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)
        self.update_status_bar()

    def sort_column(self, col, reverse):
        l = [(self.tree.set(k, col), k) for k in self.tree.get_children('')]
        try:
            l.sort(key=lambda t: float(t[0]) if t[0].replace('.', '').isdigit() else t[0].lower(), reverse=reverse)
        except ValueError:
            l.sort(key=lambda t: t[0].lower(), reverse=reverse)
        for index, (val, k) in enumerate(l):
            self.tree.move(k, '', index)
        self.tree.heading(col, command=lambda: self.sort_column(col, not reverse))

    def update_status_bar(self):
        total_cameras = len(self.cameras)
        total_models = len(self.models)
        self.status_bar.config(text=f"–í—Å–µ–≥–æ –∫–∞–º–µ—Ä: {total_cameras} | –í—Å–µ–≥–æ –º–æ–¥–µ–ª–µ–π: {total_models}")

    def backup_main_json(self):
        main_path = get_main_json_path(self.data_dir)
        if not os.path.exists(main_path):
            messagebox.showerror("–û—à–∏–±–∫–∞", "–§–∞–π–ª main.json –Ω–µ –Ω–∞–π–¥–µ–Ω!")
            return
        backup_dir = os.path.dirname(main_path)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_name = f"main_backup_{timestamp}.json"
        backup_path = os.path.join(backup_dir, backup_name)
        try:
            shutil.copy2(main_path, backup_path)
            messagebox.showinfo("–†–µ–∑–µ—Ä–≤–Ω–∞—è –∫–æ–ø–∏—è", f"–†–µ–∑–µ—Ä–≤–Ω–∞—è –∫–æ–ø–∏—è —Å–æ–∑–¥–∞–Ω–∞:\n{backup_path}")
        except Exception as e:
            messagebox.showerror("–û—à–∏–±–∫–∞", f"–ù–µ —É–¥–∞–ª–æ—Å—å —Å–æ–∑–¥–∞—Ç—å —Ä–µ–∑–µ—Ä–≤–Ω—É—é –∫–æ–ø–∏—é:\n{str(e)}")

    def open_add_camera_tab(self):
        for i in range(self.notebook.index("end") - 1, 0, -1):
            self.notebook.forget(i)
        add_frame = tk.Frame(self.notebook, padx=20, pady=20)
        self.notebook.add(add_frame, text="‚ûï –ù–æ–≤–∞—è –∫–∞–º–µ—Ä–∞")
        self.notebook.select(add_frame)

        fields = [
            ("–õ–∏–Ω–∏—è", "line", "–ù–æ–≤–∞—è –ª–∏–Ω–∏—è"),
            ("IP-–∞–¥—Ä–µ—Å", "ip", ""),
            ("–õ–æ–≥–∏–Ω Web", "web_user", ""),
            ("–ü–∞—Ä–æ–ª—å Web", "web_pass", ""),
            ("–õ–æ–≥–∏–Ω –ø–æ—Ç–æ–∫–∞", "stream_user", "admin"),
            ("–ü–∞—Ä–æ–ª—å –ø–æ—Ç–æ–∫–∞", "stream_pass", "admin"),
            ("–°–µ—Ä–∏–π–Ω—ã–π –Ω–æ–º–µ—Ä", "sn", ""),
            ("MAC-–∞–¥—Ä–µ—Å", "mac", ""),
            ("–ü—Ä–æ—à–∏–≤–∫–∞", "fw", ""),
            ("–ö–æ–º–º–µ–Ω—Ç–∞—Ä–∏–π", "comment", "")
        ]
        entries = {}
        row = 0
        for label_text, key, default in fields:
            tk.Label(add_frame, text=f"{label_text}:", font=LARGE_FONT, anchor="w").grid(row=row, column=0, sticky="w", pady=5)
            var = tk.StringVar(value=default)
            entry = tk.Entry(add_frame, textvariable=var, width=40, font=LARGE_FONT)
            entry.grid(row=row, column=1, sticky="w", padx=10, pady=5)
            entries[key] = var
            # –£–Ω–∏–≤–µ—Ä—Å–∞–ª—å–Ω–∞—è –ø—Ä–∏–≤—è–∑–∫–∞ –¥–ª—è Ctrl+C/V –Ω–∞ —Ä—É—Å—Å–∫–æ–π –∏ –∞–Ω–≥–ª–∏–π—Å–∫–æ–π —Ä–∞—Å–∫–ª–∞–¥–∫–µ
            def handle_ctrl_c_v(event, entry_ref=entry):
                if event.state & 0x4:  # Ctrl –Ω–∞–∂–∞—Ç
                    if event.keycode == 54:  # –ö–ª–∞–≤–∏—à–∞ 'C' / '–°'
                        entry_ref.event_generate("<<Copy>>")
                        return "break"
                    elif event.keycode == 58:  # –ö–ª–∞–≤–∏—à–∞ 'V' / '–ú'
                        entry_ref.event_generate("<<Paste>>")
                        return "break"
            entry.bind("<Key>", handle_ctrl_c_v)
            row += 1

        tk.Label(add_frame, text="–ú–æ–¥–µ–ª—å:", font=LARGE_FONT, anchor="w").grid(row=row, column=0, sticky="w", pady=5)
        model_var = tk.StringVar()
        model_combo = ttk.Combobox(add_frame, textvariable=model_var, values=list(self.models.keys()), state="readonly", width=38, font=LARGE_FONT)
        model_combo.grid(row=row, column=1, sticky="w", padx=10, pady=5)
        entries["model"] = model_var
        row += 1

        btn_frame = tk.Frame(add_frame)
        btn_frame.grid(row=row, column=0, columnspan=2, pady=20)
        def save_new_camera():
            new_cam = {key: var.get().strip() for key, var in entries.items()}
            if not new_cam["model"]:
                messagebox.showwarning("–û—à–∏–±–∫–∞", "–í—ã–±–µ—Ä–∏—Ç–µ –º–æ–¥–µ–ª—å –∫–∞–º–µ—Ä—ã")
                return
            self.cameras.append(new_cam)
            save_cameras(self.cameras, self.data_dir)
            self.filtered_cameras = self.cameras.copy()
            self.refresh_table()
            self.update_status_bar()
            self.show_camera_list()
            messagebox.showinfo("–£—Å–ø–µ—Ö", "–ö–∞–º–µ—Ä–∞ –¥–æ–±–∞–≤–ª–µ–Ω–∞")

        def cancel():
            self.show_camera_list()

        tk.Button(btn_frame, text="üíæ –°–æ—Ö—Ä–∞–Ω–∏—Ç—å", command=save_new_camera, font=LARGE_FONT, bg="#a0f0a0").pack(side=tk.LEFT, padx=10)
        tk.Button(btn_frame, text="‚ùå –û—Ç–º–µ–Ω–∞", command=cancel, font=LARGE_FONT).pack(side=tk.LEFT, padx=10)

    def show_camera_list(self):
        self.notebook.select(self.list_frame)

    def open_settings_tab(self):
        for i in range(self.notebook.index("end") - 1, 0, -1):
            self.notebook.forget(i)
        settings_frame = tk.Frame(self.notebook, padx=30, pady=30)
        self.notebook.add(settings_frame, text="‚öôÔ∏è –ù–∞—Å—Ç—Ä–æ–π–∫–∏")
        self.notebook.select(settings_frame)

        tk.Label(settings_frame, text="–ö–∞—Ç–∞–ª–æ–≥ —Ö—Ä–∞–Ω–µ–Ω–∏—è –¥–∞–Ω–Ω—ã—Ö:", font=BOLD_FONT).pack(anchor="w", pady=(0,10))
        path_var = tk.StringVar(value=self.settings["data_dir_path"])
        path_entry = tk.Entry(settings_frame, textvariable=path_var, width=80, font=LARGE_FONT, state="readonly")
        path_entry.pack(fill=tk.X, pady=5)

        status_label = tk.Label(settings_frame, text="", font=LARGE_FONT, fg="green")
        status_label.pack(pady=5)

        def update_status():
            path = path_var.get()
            if os.path.exists(path):
                main_path = get_main_json_path(path)
                models_path = get_models_json_path(path)
                main_exists = os.path.exists(main_path)
                models_exists = os.path.exists(models_path)
                if main_exists and models_exists:
                    status_label.config(text="‚úÖ –ö–∞—Ç–∞–ª–æ–≥ –≤–∞–ª–∏–¥–µ–Ω. –§–∞–π–ª—ã –¥–∞–Ω–Ω—ã—Ö –Ω–∞–π–¥–µ–Ω—ã.", fg="green")
                elif main_exists or models_exists:
                    status_label.config(text="‚ö†Ô∏è –ù–∞–π–¥–µ–Ω —Ç–æ–ª—å–∫–æ –æ–¥–∏–Ω —Ñ–∞–π–ª –¥–∞–Ω–Ω—ã—Ö.", fg="orange")
                else:
                    status_label.config(text="‚ö†Ô∏è –§–∞–π–ª—ã –¥–∞–Ω–Ω—ã—Ö –Ω–µ –Ω–∞–π–¥–µ–Ω—ã. –ë—É–¥—É—Ç —Å–æ–∑–¥–∞–Ω—ã –ø—Ä–∏ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏–∏.", fg="orange")
            else:
                status_label.config(text="‚ùå –ö–∞—Ç–∞–ª–æ–≥ –Ω–µ —Å—É—â–µ—Å—Ç–≤—É–µ—Ç", fg="red")

        update_status()

        def choose_new_data_dir():
            new_dir = filedialog.askdirectory(title="–í—ã–±–µ—Ä–∏—Ç–µ –∫–∞—Ç–∞–ª–æ–≥ –¥–ª—è —Ö—Ä–∞–Ω–µ–Ω–∏—è –¥–∞–Ω–Ω—ã—Ö")
            if new_dir:
                path_var.set(new_dir)
                self.settings["data_dir_path"] = new_dir
                save_settings(self.settings)
                created = ensure_data_files_exist(new_dir)
                if created:
                    msg = "–°–æ–∑–¥–∞–Ω—ã –Ω–æ–≤—ã–µ —Ñ–∞–π–ª—ã –¥–∞–Ω–Ω—ã—Ö:\n" + "\n".join(created)
                    messagebox.showinfo("–ò–Ω–∏—Ü–∏–∞–ª–∏–∑–∞—Ü–∏—è", msg)
                self.data_dir = new_dir
                self.models = load_models(self.data_dir)
                self.cameras = load_cameras(self.data_dir)
                self.filtered_cameras = self.cameras.copy()
                self.refresh_table()
                self.update_status_bar()
                update_status()
                messagebox.showinfo("–£—Å–ø–µ—Ö", "–ö–∞—Ç–∞–ª–æ–≥ –¥–∞–Ω–Ω—ã—Ö –æ–±–Ω–æ–≤–ª–µ–Ω")

        tk.Button(settings_frame, text="üìÇ –í—ã–±—Ä–∞—Ç—å –∫–∞—Ç–∞–ª–æ–≥ –¥–∞–Ω–Ω—ã—Ö", command=choose_new_data_dir, font=LARGE_FONT, bg="#d0d0d0").pack(pady=20)

        tk.Label(settings_frame, text="–ò–Ω—Ñ–æ—Ä–º–∞—Ü–∏—è:", font=BOLD_FONT).pack(anchor="w", pady=(30,10))
        tk.Label(settings_frame, text="‚Ä¢ –ü—Ä–æ–≥—Ä–∞–º–º–∞ —Ö—Ä–∞–Ω–∏—Ç –≤—Å–µ –¥–∞–Ω–Ω—ã–µ (main.json –∏ models.json) –≤ –≤—ã–±—Ä–∞–Ω–Ω–æ–º –∫–∞—Ç–∞–ª–æ–≥–µ.", font=LARGE_FONT, anchor="w").pack(anchor="w")
        tk.Label(settings_frame, text="‚Ä¢ –ü—Ä–∏ —Å–º–µ–Ω–µ –∫–∞—Ç–∞–ª–æ–≥–∞ –≤—Å–µ —Ç–µ–∫—É—â–∏–µ –∏–∑–º–µ–Ω–µ–Ω–∏—è –±—É–¥—É—Ç –ø–æ—Ç–µ—Ä—è–Ω—ã.", font=LARGE_FONT, anchor="w").pack(anchor="w")
        tk.Label(settings_frame, text="‚Ä¢ –í—Å–µ–≥–¥–∞ –¥–µ–ª–∞–π—Ç–µ —Ä–µ–∑–µ—Ä–≤–Ω—É—é –∫–æ–ø–∏—é –ø–µ—Ä–µ–¥ —Å–º–µ–Ω–æ–π –∫–∞—Ç–∞–ª–æ–≥–∞.", font=LARGE_FONT, anchor="w").pack(anchor="w")

    def open_models_tab(self):
        for i in range(self.notebook.index("end") - 1, 0, -1):
            self.notebook.forget(i)
        models_frame = tk.Frame(self.notebook, padx=20, pady=20)
        self.notebook.add(models_frame, text="üìã –ú–æ–¥–µ–ª–∏")
        self.notebook.select(models_frame)

        left_frame = tk.Frame(models_frame)
        left_frame.pack(side=tk.LEFT, fill=tk.Y, padx=(0,20))
        right_frame = tk.Frame(models_frame)
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        tk.Label(left_frame, text="–í—ã–±–µ—Ä–∏—Ç–µ –º–æ–¥–µ–ª—å:", font=BOLD_FONT).pack(anchor="w", pady=(0,10))
        self.model_listbox = tk.Listbox(left_frame, width=40, height=25, font=LARGE_FONT)
        self.model_listbox.pack()
        for model in self.models.keys():
            self.model_listbox.insert(tk.END, model)

        tk.Label(right_frame, text="–î–µ—Ç–∞–ª–∏ –º–æ–¥–µ–ª–∏:", font=BOLD_FONT).pack(anchor="w", pady=(0,10))
        tk.Label(right_frame, text="–ù–∞–∑–≤–∞–Ω–∏–µ –º–æ–¥–µ–ª–∏:", font=LARGE_FONT).pack(anchor="w")
        self.model_name_var = tk.StringVar()
        name_entry = tk.Entry(right_frame, textvariable=self.model_name_var, width=50, font=LARGE_FONT)
        name_entry.pack(anchor="w", pady=(0,10))

        tk.Label(right_frame, text="–®–∞–±–ª–æ–Ω RTSP-—Å—Å—ã–ª–∫–∏ (–æ—Å–Ω–æ–≤–Ω–æ–π):", font=LARGE_FONT).pack(anchor="w")
        self.rtsp_template_var = tk.StringVar()
        rtsp_entry = tk.Entry(right_frame, textvariable=self.rtsp_template_var, width=70, font=LARGE_FONT)
        rtsp_entry.pack(anchor="w", pady=(0,10))

        tk.Label(right_frame, text="–®–∞–±–ª–æ–Ω RTSP-—Å—Å—ã–ª–∫–∏ (–¥–æ–ø. 1):", font=LARGE_FONT).pack(anchor="w")
        self.rtsp_template_2_var = tk.StringVar()
        rtsp_entry_2 = tk.Entry(right_frame, textvariable=self.rtsp_template_2_var, width=70, font=LARGE_FONT)
        rtsp_entry_2.pack(anchor="w", pady=(0,10))

        tk.Label(right_frame, text="–®–∞–±–ª–æ–Ω RTSP-—Å—Å—ã–ª–∫–∏ (–¥–æ–ø. 2):", font=LARGE_FONT).pack(anchor="w")
        self.rtsp_template_3_var = tk.StringVar()
        rtsp_entry_3 = tk.Entry(right_frame, textvariable=self.rtsp_template_3_var, width=70, font=LARGE_FONT)
        rtsp_entry_3.pack(anchor="w", pady=(0,10))

        tk.Label(right_frame, text="–†–∞–∑—Ä–µ—à–µ–Ω–∏–µ:", font=LARGE_FONT).pack(anchor="w")
        self.resolution_var = tk.StringVar()
        resolution_entry = tk.Entry(right_frame, textvariable=self.resolution_var, width=20, font=LARGE_FONT)
        resolution_entry.pack(anchor="w", pady=(0,10))

        tk.Label(right_frame, text="–ö–æ—Ä–ø—É—Å:", font=LARGE_FONT).pack(anchor="w")
        self.housing_var = tk.StringVar()
        housing_entry = tk.Entry(right_frame, textvariable=self.housing_var, width=70, font=LARGE_FONT)
        housing_entry.pack(anchor="w", pady=(0,10))

        tk.Label(right_frame, text="–ë—Ä–∞—É–∑–µ—Ä:", font=LARGE_FONT).pack(anchor="w")
        self.browser_var = tk.StringVar()
        browser_entry = tk.Entry(right_frame, textvariable=self.browser_var, width=70, font=LARGE_FONT)
        browser_entry.pack(anchor="w", pady=(0,10))

        tk.Label(right_frame, text="–ü—Ä–∏–º–µ—á–∞–Ω–∏–µ:", font=LARGE_FONT).pack(anchor="w")
        self.note_var = tk.StringVar()
        note_entry = tk.Entry(right_frame, textvariable=self.note_var, width=70, font=LARGE_FONT)
        note_entry.pack(anchor="w", pady=(0,20))

        btn_frame = tk.Frame(right_frame)
        btn_frame.pack()

        def on_model_select(event):
            selection = self.model_listbox.curselection()
            if not selection:
                return
            model_name = self.model_listbox.get(selection[0])
            model_data = self.models[model_name]
            self.model_name_var.set(model_name)
            self.rtsp_template_var.set(model_data.get("rtsp_template", ""))
            self.rtsp_template_2_var.set(model_data.get("rtsp_template_2", ""))
            self.rtsp_template_3_var.set(model_data.get("rtsp_template_3", ""))
            self.resolution_var.set(model_data.get("resolution", ""))
            self.housing_var.set(model_data.get("housing", ""))
            self.browser_var.set(model_data.get("browser", ""))
            self.note_var.set(model_data.get("note", ""))

        def save_model():
            selection = self.model_listbox.curselection()
            old_name = self.model_listbox.get(selection[0]) if selection else ""
            new_name = self.model_name_var.get().strip()
            if not new_name:
                messagebox.showwarning("–û—à–∏–±–∫–∞", "–ò–º—è –º–æ–¥–µ–ª–∏ –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—É—Å—Ç—ã–º")
                return
            if old_name and old_name != new_name:
                if old_name in self.models:
                    del self.models[old_name]
            self.models[new_name] = {
                "rtsp_template": self.rtsp_template_var.get(),
                "rtsp_template_2": self.rtsp_template_2_var.get(),
                "rtsp_template_3": self.rtsp_template_3_var.get(),
                "resolution": self.resolution_var.get(),
                "housing": self.housing_var.get(),
                "browser": self.browser_var.get(),
                "note": self.note_var.get()
            }
            save_models(self.models, self.data_dir)
            self.refresh_model_list()
            self.update_status_bar()
            messagebox.showinfo("–£—Å–ø–µ—Ö", "–ú–æ–¥–µ–ª—å —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∞")

        def add_new_model():
            self.model_listbox.selection_clear(0, tk.END)
            self.model_name_var.set("")
            self.rtsp_template_var.set("")
            self.rtsp_template_2_var.set("")
            self.rtsp_template_3_var.set("")
            self.resolution_var.set("")
            self.housing_var.set("")
            self.browser_var.set("")
            self.note_var.set("")

        def delete_model():
            selection = self.model_listbox.curselection()
            if not selection:
                messagebox.showwarning("–û—à–∏–±–∫–∞", "–°–Ω–∞—á–∞–ª–∞ –≤—ã–±–µ—Ä–∏—Ç–µ –º–æ–¥–µ–ª—å –¥–ª—è —É–¥–∞–ª–µ–Ω–∏—è")
                return
            model_name = self.model_listbox.get(selection[0])
            if messagebox.askyesno("–ü–æ–¥—Ç–≤–µ—Ä–∂–¥–µ–Ω–∏–µ", f"–£–¥–∞–ª–∏—Ç—å –º–æ–¥–µ–ª—å '{model_name}'?"):
                if model_name in self.models:
                    del self.models[model_name]
                    save_models(self.models, self.data_dir)
                    self.refresh_model_list()
                    self.update_status_bar()
                    add_new_model()
                    messagebox.showinfo("–£—Å–ø–µ—Ö", "–ú–æ–¥–µ–ª—å —É–¥–∞–ª–µ–Ω–∞")

        tk.Button(btn_frame, text="üíæ –°–æ—Ö—Ä–∞–Ω–∏—Ç—å", command=save_model, font=LARGE_FONT, bg="#a0f0a0").pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="‚ûï –î–æ–±–∞–≤–∏—Ç—å –Ω–æ–≤—É—é", command=add_new_model, font=LARGE_FONT).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="üóëÔ∏è –£–¥–∞–ª–∏—Ç—å", command=delete_model, font=LARGE_FONT).pack(side=tk.LEFT, padx=5)

        self.model_listbox.bind("<<ListboxSelect>>", on_model_select)

    def refresh_model_list(self):
        self.model_listbox.delete(0, tk.END)
        for model in self.models.keys():
            self.model_listbox.insert(tk.END, model)

    def schedule_button_placement(self, event=None):
        if hasattr(self, '_after_id'):
            self.root.after_cancel(self._after_id)
        self._after_id = self.root.after(50, self.place_action_buttons)

    def place_action_buttons(self):
        for btn in self.rtsp_buttons:
            btn.destroy()
        self.rtsp_buttons.clear()
        for child in self.tree.get_children():
            bbox = self.tree.bbox(child, column="rtsp_btn")
            if not bbox:
                continue
            x, y, width, height = bbox
            # –ö–Ω–æ–ø–∫–∞ RTSP –∑–∞–Ω–∏–º–∞–µ—Ç –≤—Å—é —è—á–µ–π–∫—É
            rtsp_btn = ttk.Button(self.tree, text="RTSP", width=5, command=lambda c=child: self.open_rtsp_detail(c))
            rtsp_btn.place(x=x+2, y=y+2, width=width-4, height=height-4)
            self.rtsp_buttons.append(rtsp_btn)

    def open_rtsp_detail(self, item_id):
        idx = int(item_id)
        if idx >= len(self.filtered_cameras):
            return
        filtered_cam = self.filtered_cameras[idx]
        real_idx = None
        for i, cam in enumerate(self.cameras):
            if (cam.get("line") == filtered_cam.get("line") and
                cam.get("ip") == filtered_cam.get("ip") and
                cam.get("model") == filtered_cam.get("model")):
                real_idx = i
                break
        if real_idx is not None:
            RTSPDetailWindow(self.root, real_idx, self.cameras, self.models, self, self.data_dir)

    def on_search_change(self, *_):
        query = self.search_var.get().lower().strip()
        if not query:
            self.filtered_cameras = self.cameras.copy()
        else:
            self.filtered_cameras = [
                cam for cam in self.cameras
                if query in cam.get("line", "").lower() or
                   query in cam.get("model", "").lower() or
                   query in cam.get("ip", "").lower()
            ]
        self.refresh_table()

    def refresh_table(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        for btn in self.rtsp_buttons:
            btn.destroy()
        self.rtsp_buttons.clear()
        for idx, cam in enumerate(self.filtered_cameras):
            values = (
                cam.get("line", ""),
                cam.get("model", ""),
                cam.get("ip", ""),
                cam.get("web_user", ""),
                cam.get("web_pass", ""),
                cam.get("stream_user", ""),
                cam.get("stream_pass", ""),
                cam.get("sn", ""),
                cam.get("mac", ""),
                cam.get("fw", ""),
                cam.get("comment", ""),
                ""
            )
            self.tree.insert("", "end", iid=str(idx), values=values)
        self.schedule_button_placement()

    def on_double_click_edit(self, event):
        region = self.tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        item = self.tree.identify_row(event.y)
        column = self.tree.identify_column(event.x)
        if not item or not column:
            return
        col_index = int(column[1:]) - 1
        columns = ("line", "model", "ip", "web_user", "web_pass",
                   "stream_user", "stream_pass", "sn", "mac", "fw", "comment")
        if col_index >= len(columns) or col_index < 0:
            return
        column_name = columns[col_index]
        current_value = self.tree.set(item, column)
        if column_name == "model":
            self.create_model_selector_window(item, current_value)
        else:
            self.create_edit_window(item, column_name, current_value)

    def create_model_selector_window(self, item, current_value):
        def on_model_selected(new_value):
            filtered_idx = int(item)
            if filtered_idx < len(self.filtered_cameras):
                filtered_cam = self.filtered_cameras[filtered_idx]
                for i, cam in enumerate(self.cameras):
                    if (cam.get("line") == filtered_cam.get("line") and
                            cam.get("ip") == filtered_cam.get("ip") and
                            cam.get("model") == filtered_cam.get("model")):
                        self.cameras[i]["model"] = new_value
                        break
                self.filtered_cameras[filtered_idx]["model"] = new_value
                save_cameras(self.cameras, self.data_dir)
                self.refresh_table()
        ModelSelectorWindow(self.root, current_value, list(self.models.keys()), on_model_selected)

    def create_edit_window(self, item, column_name, current_value):
        edit_win = tk.Toplevel(self.root)
        edit_win.title(f"–†–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ: {column_name}")
        edit_win.geometry("400x150")
        edit_win.transient(self.root)
        edit_win.grab_set()
        center_window(edit_win, self.root)
        tk.Label(edit_win, text=f"–†–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ {column_name}:", font=LARGE_FONT).pack(pady=10)
        entry_var = tk.StringVar(value=current_value)
        entry = tk.Entry(edit_win, textvariable=entry_var, font=LARGE_FONT, width=40)
        entry.pack(pady=5, padx=20)
        # –£–Ω–∏–≤–µ—Ä—Å–∞–ª—å–Ω–∞—è –ø—Ä–∏–≤—è–∑–∫–∞ –¥–ª—è Ctrl+C/V –Ω–∞ —Ä—É—Å—Å–∫–æ–π –∏ –∞–Ω–≥–ª–∏–π—Å–∫–æ–π —Ä–∞—Å–∫–ª–∞–¥–∫–µ
        def handle_ctrl_c_v(event):
            if event.state & 0x4:  # Ctrl –Ω–∞–∂–∞—Ç
                if event.keycode == 54:  # –ö–ª–∞–≤–∏—à–∞ 'C' / '–°'
                    entry.event_generate("<<Copy>>")
                    return "break"
                elif event.keycode == 58:  # –ö–ª–∞–≤–∏—à–∞ 'V' / '–ú'
                    entry.event_generate("<<Paste>>")
                    return "break"
        entry.bind("<Key>", handle_ctrl_c_v)
        entry.select_range(0, tk.END)
        entry.focus_set()
        def save_edit():
            new_value = entry_var.get().strip()
            if new_value != current_value:
                filtered_idx = int(item)
                if filtered_idx < len(self.filtered_cameras):
                    filtered_cam = self.filtered_cameras[filtered_idx]
                    for i, cam in enumerate(self.cameras):
                        if (cam.get("line") == filtered_cam.get("line") and
                                cam.get("ip") == filtered_cam.get("ip") and
                                cam.get("model") == filtered_cam.get("model")):
                            self.cameras[i][column_name] = new_value
                            break
                    self.filtered_cameras[filtered_idx][column_name] = new_value
                    save_cameras(self.cameras, self.data_dir)
                    self.refresh_table()
            edit_win.destroy()
        def cancel_edit():
            edit_win.destroy()
        btn_frame = tk.Frame(edit_win)
        btn_frame.pack(pady=10)
        tk.Button(btn_frame, text="üíæ –°–æ—Ö—Ä–∞–Ω–∏—Ç—å", command=save_edit, font=LARGE_FONT, bg="#a0f0a0").pack(side=tk.LEFT, padx=10)
        tk.Button(btn_frame, text="‚ùå –û—Ç–º–µ–Ω–∞", command=cancel_edit, font=LARGE_FONT).pack(side=tk.LEFT, padx=10)
        entry.bind("<Return>", lambda e: save_edit())
        entry.bind("<Escape>", lambda e: cancel_edit())

    def on_right_click(self, event):
        item = self.tree.identify_row(event.y)
        if not item:
            return
        filtered_idx = int(item)
        if filtered_idx >= len(self.filtered_cameras):
            return
        filtered_camera = self.filtered_cameras[filtered_idx]
        main_idx = None
        for i, camera in enumerate(self.cameras):
            if (camera.get("line") == filtered_camera.get("line") and
                    camera.get("ip") == filtered_camera.get("ip") and
                    camera.get("model") == filtered_camera.get("model")):
                main_idx = i
                break
        if main_idx is not None:
            menu = tk.Menu(self.root, tearoff=0, font=LARGE_FONT)
            menu.add_command(label="üóëÔ∏è –£–¥–∞–ª–∏—Ç—å –∑–∞–ø–∏—Å—å", command=lambda: self.delete_camera(main_idx, filtered_idx))
            menu.post(event.x_root, event.y_root)

    def delete_camera(self, main_idx, filtered_idx):
        camera = self.cameras[main_idx]
        camera_info = f"{camera.get('line', '')} - {camera.get('model', '')} - {camera.get('ip', '')}"
        if messagebox.askyesno("–ü–æ–¥—Ç–≤–µ—Ä–∂–¥–µ–Ω–∏–µ —É–¥–∞–ª–µ–Ω–∏—è", f"–í—ã —É–≤–µ—Ä–µ–Ω—ã, —á—Ç–æ —Ö–æ—Ç–∏—Ç–µ —É–¥–∞–ª–∏—Ç—å –∫–∞–º–µ—Ä—É:\n{camera_info}?"):
            del self.cameras[main_idx]
            save_cameras(self.cameras, self.data_dir)
            self.filtered_cameras = [cam for cam in self.cameras
                                     if self.search_var.get().lower() in cam.get("line", "").lower() or
                                     self.search_var.get().lower() in cam.get("model", "").lower() or
                                     self.search_var.get().lower() in cam.get("ip", "").lower()]
            self.refresh_table()
            self.update_status_bar()
            messagebox.showinfo("–£—Å–ø–µ—Ö", "–ó–∞–ø–∏—Å—å —É–¥–∞–ª–µ–Ω–∞")

    def export_excel(self):
        filepath = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if filepath:
            export_to_excel(self.cameras, filepath)

    def export_pdf(self):
        filepath = filedialog.asksaveasfilename(defaultextension=".pdf",
                                                filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")])
        if filepath:
            loading_win = LoadingWindow(self.root, "–≠–∫—Å–ø–æ—Ä—Ç PDF", "–ò–¥–µ—Ç —Å–æ–∑–¥–∞–Ω–∏–µ PDF-–¥–æ–∫—É–º–µ–Ω—Ç–∞...")
            self.root.update()
            try:
                export_to_pdf(self.cameras, self.data_dir, filepath)
            finally:
                loading_win.destroy()

    def generate_gallery(self):
        filepath = filedialog.asksaveasfilename(defaultextension=".jpg",
                                                filetypes=[("JPEG files", "*.jpg"), ("PNG files", "*.png"), ("All files", "*.*")])
        if filepath:
            loading_win = LoadingWindow(self.root, "–°–æ–∑–¥–∞–Ω–∏–µ –≥–∞–ª–µ—Ä–µ–∏", "–ò–¥–µ—Ç —Ñ–æ—Ä–º–∏—Ä–æ–≤–∞–Ω–∏–µ –≥–∞–ª–µ—Ä–µ–∏ —Å–∫—Ä–∏–Ω—à–æ—Ç–æ–≤...")
            self.root.update()
            try:
                generate_gallery(self.cameras, self.data_dir, filepath)
            finally:
                loading_win.destroy()

    def show_info(self):
        info_win = tk.Toplevel(self.root)
        info_win.title("–ò–Ω—Ñ–æ—Ä–º–∞—Ü–∏—è –æ –ø—Ä–æ–≥—Ä–∞–º–º–µ")
        info_win.geometry("400x300")
        info_win.transient(self.root)
        info_win.grab_set()
        center_window(info_win, self.root)
        ico_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ico.jpg")
        try:
            pil_image = Image.open(ico_path)
            pil_image.thumbnail((120, 120), Image.Resampling.LANCZOS)
            photo = ImageTk.PhotoImage(pil_image)
            img_label = tk.Label(info_win, image=photo)
            img_label.image = photo
            img_label.pack(pady=15)
        except Exception as e:
            tk.Label(info_win, text="–õ–æ–≥–æ—Ç–∏–ø", font=("Arial", 16, "bold")).pack(pady=15)
            print(f"–ù–µ —É–¥–∞–ª–æ—Å—å –∑–∞–≥—Ä—É–∑–∏—Ç—å ico.jpg: {e}")
        info_text = "–ü—Ä–æ–≥—Ä–∞–º–º–∞ \"Datacam\"\n–í–µ—Ä—Å–∏—è 0.4\n–ê–≤—Ç–æ—Ä: –†–∞–∑–∏–Ω –ì.–í.\n¬© 2025"
        tk.Label(info_win, text=info_text, font=("Arial", 14), justify=tk.CENTER).pack(pady=10)
        tk.Button(info_win, text="–ó–∞–∫—Ä—ã—Ç—å", command=info_win.destroy, font=LARGE_FONT, width=15).pack(pady=20)

if __name__ == "__main__":
    root = tk.Tk()
    app = CameraApp(root)
    root.mainloop()